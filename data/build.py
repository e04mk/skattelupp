#!/usr/bin/env python3
"""
Build tax-data.json for Skattelupp from official Skatteverket + SCB sources.
See SOURCES.md for provenance of every figure.
"""
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent
RAW = BASE / "raw"

# ---------------------------------------------------------------------------
# 1. Skatteverket kommunal/landstings tax rates 2026
# ---------------------------------------------------------------------------

def load_tax_rates():
    wb = openpyxl.load_workbook(RAW / "skattesatser-kommuner-2026.xlsx", data_only=True)
    ws = wb["Skattesatser kommuner 2026"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    kommunal = {}   # kommunkod -> rate (fraction)
    landsting = {}  # lankod -> rate (fraction)
    kommun_name_titlecase = {}  # kommunkod -> name as printed (ALLCAPS) - fallback only

    for r in rows:
        year, kod, kommun, forsamling, summa_inkl, summa_exkl, kommunal_skatt, landstings_skatt, begr, kyrko = r
        kod_clean = kod.replace(" ", "")
        lan_code = kod_clean[:2]
        kommunal[kod_clean] = round(float(kommunal_skatt) / 100.0, 6)
        landsting[lan_code] = round(float(landstings_skatt) / 100.0, 6)
        kommun_name_titlecase[kod_clean] = kommun

    assert len(kommunal) == 290, f"expected 290 kommuner, got {len(kommunal)}"
    assert len(landsting) == 21, f"expected 21 lan, got {len(landsting)}"
    return kommunal, landsting, kommun_name_titlecase


# ---------------------------------------------------------------------------
# 2. Canonical kommun / region names from SCB table metadata
# ---------------------------------------------------------------------------

def load_kommun_names():
    d = json.load(open(RAW / "tab4199_meta.json"))
    cat = d["dimension"]["Region"]["category"]
    labels = cat["label"]
    names = {}
    for code, label in labels.items():
        if code == "00":
            continue
        names[code] = label
    assert len(names) == 290
    return names


REGION_NAME_OVERRIDES = {
    "14": "Västra Götaland",
}


def clean_region_name(raw_label):
    # raw examples: "Region Stockholm", "Region Jönköpings län", "Västra Götalandsregionen",
    # "Region Kalmar län", "Region Gotland", "Region Jämtland Härjedalen"
    name = raw_label
    name = re.sub(r"^Region\s+", "", name)
    name = re.sub(r"regionen$", "", name)  # Västra Götalandsregionen -> Västra Götaland
    name = re.sub(r"s län$", "", name)
    name = re.sub(r"\s+län$", "", name)
    return name.strip()


def load_region_names():
    d = json.load(open(RAW / "tab4242_meta.json"))
    cat = d["dimension"]["Region"]["category"]
    labels = cat["label"]
    names = {}
    for code, label in labels.items():
        if code == "00":
            continue
        m = re.match(r"^(\d+)L$", code)
        assert m, f"unexpected region code {code}"
        digits = m.group(1)
        lan_code = digits[:2]
        clean = clean_region_name(label)
        clean = REGION_NAME_OVERRIDES.get(lan_code, clean)
        names[lan_code] = clean
    assert len(names) == 21
    return names


# ---------------------------------------------------------------------------
# 3. SCB RS (Räkenskapssammandraget) net-cost data -> category shares
# ---------------------------------------------------------------------------

KOMMUN_CATEGORIES = {
    "Förskola & grundskola": ["400", "407", "412", "415", "425", "435", "440", "443"],
    "Gymnasieskola & vuxenutbildning": ["450", "453", "470", "472", "474", "475", "476", "478"],
    "Äldreomsorg": ["500", "505", "510"],
    "Individ- och familjeomsorg / stöd till personer med funktionsnedsättning": [
        "520", "513", "530", "559", "569", "571", "575", "585",
    ],
    # These four used to be the pre-aggregated SCB rollup codes (290/390/190/690+890);
    # switched to their constituent leaf codes (verified to sum to the same rollup
    # totals) so the same fetch also powers the "Visa som tabell" drill-down below.
    "Infrastruktur, skydd & miljö": [
        "215", "220", "225", "230", "249", "250", "261", "263", "267", "270", "275",
    ],
    "Kultur & fritid": ["310", "315", "320", "330", "300", "340", "350"],
    "Politisk verksamhet & administration": ["100", "110", "120", "130"],
    "Övrigt": ["600", "610", "800", "805", "810", "815", "830", "832", "834", "855", "860", "865", "870"],
}

REGION_CATEGORIES = {
    "Hälso- och sjukvård": ["0", "1", "2", "4"],
    "Tandvård": ["3"],
    "Kollektivtrafik": ["7"],
    "Regional utveckling": ["5", "6", "8"],
    "Politisk verksamhet & administration": ["910", "920"],
    # Was the "940-980" rollup code + "Jamf"; switched to the rollup's own leaves
    # (940/960/980) so the same fetch also powers the detail drill-down below.
    "Övrigt": ["940", "960", "980", "Jamf"],
}

# ---------------------------------------------------------------------------
# 3a-detail. Fine-grained sub-items within each top-level category, for the
# "Visa som tabell" drill-down. Codes/labels are straight from the SCB
# Verksomrkom dimension labels in raw/tab4199_meta.json and raw/tab4242_meta.json
# (see SOURCES.md). Categories built from a single already-atomic code (e.g.
# region "Tandvård") have no further breakdown available and are omitted here.
KOMMUN_CODE_LABELS = {
    "400": "Öppen förskola", "407": "Förskola", "412": "Pedagogisk omsorg",
    "415": "Öppen fritidsverksamhet", "425": "Fritidshem", "435": "Förskoleklass",
    "440": "Grundskola", "443": "Anpassad grundskola",
    "450": "Gymnasieskola", "453": "Anpassad gymnasieskola",
    "470": "Grundläggande vuxenutbildning", "472": "Gymnasial vuxen- och påbyggnadsutbildning",
    "474": "Komvux anpassad utbildning", "475": "Högskoleutbildning m.m.",
    "476": "Svenska för invandrare", "478": "Uppdragsutbildning m.m.",
    "500": "Primärvård (kommunal hemsjukvård)", "505": "Hälso- och sjukvård, övrigt",
    "510": "Vård och omsorg om äldre",
    "520": "Insatser till personer med funktionsnedsättning", "513": "Insatser enligt LSS/SFB",
    "530": "Färdtjänst/riksfärdtjänst", "559": "Vård för vuxna med missbruksproblem",
    "569": "Barn- och ungdomsvård", "571": "Övriga insatser till vuxna",
    "575": "Ekonomiskt bistånd", "585": "Familjerätt och familjerådgivning",
    "215": "Fysisk och teknisk planering, bostadsförbättring", "220": "Näringslivsfrämjande åtgärder",
    "225": "Konsument- och energirådgivning", "230": "Turistverksamhet",
    "249": "Väg- och järnvägsnät, parkering", "250": "Parker",
    "261": "Miljö- och hälsoskydd, myndighetsutövning", "263": "Miljö, hälsa och hållbar utveckling",
    "267": "Alkoholtillstånd m.m.", "270": "Räddningstjänst",
    "275": "Totalförsvar och samhällsskydd",
    "310": "Stöd till studieorganisationer", "315": "Allmän kulturverksamhet",
    "320": "Bibliotek", "330": "Musikskola/kulturskola", "300": "Allmän fritidsverksamhet",
    "340": "Idrotts- och fritidsanläggningar", "350": "Fritidsgårdar",
    "100": "Nämnd- och styrelseverksamhet", "110": "Stöd till politiska partier",
    "120": "Revision", "130": "Övrig politisk verksamhet",
    "600": "Flyktingmottagande", "610": "Arbetsmarknadsåtgärder",
    "800": "Arbetsområden och lokaler", "805": "Hamnverksamhet",
    "810": "Kommersiell verksamhet", "815": "Bostadsverksamhet", "830": "Flygtrafik",
    "832": "Buss, bil och spårbunden persontrafik", "834": "Sjötrafik",
    "855": "Elförsörjning och gasförsörjning", "860": "Fjärrvärmeförsörjning",
    "865": "Vattenförsörjning och avloppshantering", "870": "Avfallshantering",
}

# Every kommun category is now already defined by its leaf codes (see
# KOMMUN_CATEGORIES above), so the detail breakdown is just those same codes -
# except "Övrigt", which is excluded: its leaves are largely internal kommunal
# business-accounting entries (e.g. "Arbetsområden och lokaler", an internal
# real-estate cost-allocation line) that can swing hugely negative in ways
# individual categories elsewhere don't. Floored at 0 leaf-by-leaf, that
# produces a visible "detail" sum many times larger than the category's real
# (small) net share - e.g. Stockholm's Övrigt is really ~0.5% of its budget,
# but summing only its positive leaves gives ~3.3%. Every other category's
# detail sum matches its top-level share to within ~1 percentage point across
# all 290 kommuner; "Övrigt" alone was off by more everywhere it was checked,
# so it's left as an opaque residual instead of a misleading breakdown.
KOMMUN_CATEGORY_DETAIL_CODES = {cat: codes for cat, codes in KOMMUN_CATEGORIES.items() if cat != "Övrigt"}

REGION_CODE_LABELS = {
    "0": "Primärvård", "1": "Specialiserad somatisk vård", "2": "Specialiserad psykiatrisk vård",
    "4": "Övrig hälso- och sjukvård",
    "5": "Utbildning (region-driven)", "6": "Kultur", "8": "Allmän regional utveckling",
    "910": "Politisk verksamhet, hälso- och sjukvård", "920": "Politisk verksamhet, regional utveckling",
    "940": "Medicinsk service", "960": "Allmän service", "980": "Fastighetsförvaltning",
    "Jamf": "Jämförelsestörande poster",
}

# Same story for regions - every multi-code category is already leaf-based
# above, so detail is a straight copy, minus the two single-code categories
# (Tandvård, Kollektivtrafik) with nothing further to break down, and minus
# "Övrigt" for the same reason as the kommun side above (its codes include
# "Jamf" - "Jämförelsestörande poster", i.e. accounting adjustment items by
# definition, not a real spending line).
REGION_CATEGORY_DETAIL_CODES = {
    cat: codes for cat, codes in REGION_CATEGORIES.items() if len(codes) > 1 and cat != "Övrigt"
}


def compute_detail(table, detail_codes_by_category, code_labels, totals, exclude_codes_by_region=None):
    """Return {region_code: {category: [{"name", "share"}, ...]}}, sub-items sorted
    descending by share. Shares are of the SAME grand total used for the
    top-level category shares (from `totals`), so a category's sub-item shares
    sum back to that category's own top-level share."""
    exclude_codes_by_region = exclude_codes_by_region or {}
    out = {}
    for rcode, row in table.items():
        total = totals.get(rcode)
        if not total:
            continue
        excluded = exclude_codes_by_region.get(rcode, set())
        cat_detail = {}
        for cat, codes in detail_codes_by_category.items():
            items = []
            for code in codes:
                if code in excluded:
                    continue
                v = max(row.get(code, 0.0), 0.0)
                if v <= 0:
                    continue
                items.append({"name": code_labels[code], "share": round(v / total, 6)})
            items.sort(key=lambda x: -x["share"])
            if items:
                cat_detail[cat] = items
        out[rcode] = cat_detail
    return out

# ---------------------------------------------------------------------------
# 3b. Statens budget - utgiftsomraden (UO1-UO27) -> top-level categories
# ---------------------------------------------------------------------------
# See SOURCES.md section "Statens budget per utgiftsomrade" for exact UO names,
# amounts, and the reasoning behind this grouping.
STATE_CATEGORIES = {
    "Vård, omsorg & socialförsäkringar": [9, 10, 11, 12],
    "Arbetsmarknad, näringsliv & infrastruktur": [14, 19, 21, 22, 23, 24],
    "Utbildning & forskning": [15, 16],
    "Allmänna bidrag till kommuner och regioner": [25],
    "Rättsväsende, försvar & samhällsskydd": [4, 6],
    "Migration": [8],
    "Bistånd & internationellt": [5, 7],
    "Rikets styrelse & allmän förvaltning": [1, 2, 3],
    "Räntor på statsskulden": [26],
    "Övrigt": [13, 17, 18, 20, 27],
}
assert sorted(sum(STATE_CATEGORIES.values(), [])) == list(range(1, 28))


def load_jsonstat(path):
    d = json.load(open(path))
    dims = d["id"]
    size = d["size"]
    assert dims[:2] == ["Region", "Verksomrkom"] or dims[0] == "Region"
    region_dim = d["dimension"]["Region"]["category"]
    verk_dim = d["dimension"]["Verksomrkom"]["category"]
    r_index = region_dim["index"]
    v_index = verk_dim["index"]
    r_label = region_dim["label"]
    v_label = verk_dim["label"]
    nR, nV = size[0], size[1]
    # remaining dims (ContentsCode, Tid) are size 1 each
    vals = d["value"]
    assert len(vals) == nR * nV, (len(vals), nR, nV)
    table = {}  # region_code -> {verk_code: value}
    for rcode, ri in r_index.items():
        row = {}
        for vcode, vi in v_index.items():
            flat = ri * nV + vi
            v = vals[flat]
            row[vcode] = float(v) if v is not None else 0.0
        table[rcode] = row
    return table, r_label, v_label


GOTLAND_KOMMUN_CODE = "0980"
# Region Gotland is a single legal entity performing both kommun AND region duties, but SCB
# RS requires it to submit BOTH a kommun-template and a region-template return. As a result,
# Gotland's KOMMUN-level RS row also carries verksamhet 500 (Primarvard) and 505 (Halso- och
# sjukvard, ovrigt) - i.e. the full regional healthcare budget - duplicated from what is
# already captured, properly, in the REGION-level RS row for Region Gotland (region code 09).
# Left in, this would make ~35% of "kommun" Gotland's reported spend look like elderly-care/
# healthcare, wildly out of line with every other kommun (health care is a REGION
# responsibility everywhere else). We exclude 500/505 from Gotland's kommun-level totals so
# its category shares are comparable to the other 289 kommuner; Gotland's real healthcare
# spend is fully represented in regions["09"].spendingShares instead.
GOTLAND_EXCLUDE_CODES = {"500", "505"}


def compute_shares(table, categories, exclude_codes_by_region=None):
    """Return {region_code: {category: share}} plus raw totals for QA."""
    exclude_codes_by_region = exclude_codes_by_region or {}
    out = {}
    totals = {}
    for rcode, row in table.items():
        excluded = exclude_codes_by_region.get(rcode, set())
        cat_sums = {}
        for cat, codes in categories.items():
            s = sum(row.get(c, 0.0) for c in codes if c not in excluded)
            # A handful of kommuner/regioner show a small NEGATIVE net cost in a category
            # (almost always "Ovrigt", driven by fee-financed affarsverksamhet like
            # elforsorjning where fee income exceeded cost that year). A negative slice
            # cannot be rendered in a share-based pie chart, so we floor it at 0; the
            # (small, single-digit-percent-at-most) amount is simply dropped rather than
            # redistributed, since it represents a net GAIN, not an unaccounted-for cost.
            cat_sums[cat] = max(s, 0.0)
        total = sum(cat_sums.values())
        totals[rcode] = total
        if total <= 0:
            out[rcode] = None  # signal: no usable data
        else:
            shares = {cat: v / total for cat, v in cat_sums.items()}
            rounded = {cat: round(v, 6) for cat, v in shares.items()}
            # nudge the largest category so the rounded shares sum to exactly 1.0
            diff = round(1.0 - sum(rounded.values()), 6)
            if diff != 0:
                largest_cat = max(rounded, key=rounded.get)
                rounded[largest_cat] = round(rounded[largest_cat] + diff, 6)
            out[rcode] = rounded
    return out, totals


# ---------------------------------------------------------------------------
# 3c. Statens budget: outcome (utfall) 2024 and adopted/proposed budget 2026,
#     by utgiftsomrade -> STATE_CATEGORIES shares
# ---------------------------------------------------------------------------

def load_state_outcome_by_uo(year=2024):
    """Sum 'Utfall' (mkr) per utgiftsomrade (1-27) for the given year, from
    Statskontoret's 'Arsutfall utgifter 1997-2025, definitivt' open-data file
    (raw/arsutfall-utgifter-1997-2025.xlsx, sheet 'data'). Rows with no
    utgiftsomrade (Utgiftstak, Marginal till utgiftstaket, Alderspensions-
    systemet vid sidan av statens budget, Riksgaldskontorets nettoutlaning,
    Kassamassig korrigering, Forandring av anslagsbehallningar) are outside
    the 27 utgiftsomraden and are correctly excluded by this filter."""
    wb = openpyxl.load_workbook(RAW / "arsutfall-utgifter-1997-2025.xlsx", data_only=True)
    ws = wb["data"]
    totals = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        uo, ar, utfall = row[0], row[4], row[10]
        if ar != year or uo is None:
            continue
        totals[int(uo)] += float(utfall or 0.0)
    assert set(totals.keys()) == set(range(1, 28)), f"expected UO 1-27, got {sorted(totals)}"
    return dict(totals)


def load_state_budget_by_uo():
    """Sum the top-level utgiftsomrade rows (mkr) from the government's
    'Specifikation av budgetens utgifter for 2026' spreadsheet (raw/
    statsbudget-2026-specifikation.xlsx, sheet 'Utgifter'), the same document
    submitted to riksdagen with the budget proposition (prop. 2025/26:1).
    The sheet lists one top-level row per utgiftsomrade (column A = UO
    number, column C = total in tkr) followed by indented anslag/anslagspost
    rows (column A blank) which we skip."""
    wb = openpyxl.load_workbook(RAW / "statsbudget-2026-specifikation.xlsx", data_only=True)
    ws = wb["Utgifter"]
    totals = {}
    for row in ws.iter_rows(min_row=1, max_row=600, values_only=True):
        uo, _name, amount_tkr = row[0], row[1], row[2]
        if uo is not None and str(uo).strip().isdigit():
            totals[int(uo)] = float(amount_tkr) / 1000.0  # tkr -> mkr
    assert set(totals.keys()) == set(range(1, 28)), f"expected UO 1-27, got {sorted(totals)}"
    return totals


def load_uo_names():
    """The official Swedish name of each utgiftsomrade (1-27), read straight from
    column B of the same budget spreadsheet used by load_state_budget_by_uo -
    the authoritative source for how the government itself labels each area."""
    wb = openpyxl.load_workbook(RAW / "statsbudget-2026-specifikation.xlsx", data_only=True)
    ws = wb["Utgifter"]
    names = {}
    for row in ws.iter_rows(min_row=1, max_row=600, values_only=True):
        uo, name, _amount = row[0], row[1], row[2]
        if uo is not None and str(uo).strip().isdigit():
            names[int(uo)] = name.strip()
    assert set(names.keys()) == set(range(1, 28)), f"expected UO 1-27, got {sorted(names)}"
    return names


def compute_state_detail(uo_totals, categories, uo_names, total):
    """Per-category list of {"name", "share"} for each utgiftsomrade in that
    category, sorted descending by share - the state-side equivalent of
    compute_detail(). Shares are of the same grand `total` used for the
    top-level category shares."""
    out = {}
    for cat, uos in categories.items():
        items = [
            {"name": uo_names[u], "share": round(uo_totals[u] / total, 6)}
            for u in uos
        ]
        items.sort(key=lambda x: -x["share"])
        out[cat] = items
    return out


def compute_state_shares(uo_totals, categories):
    """Aggregate per-UO totals into category totals and normalize to shares
    summing to exactly 1.0 (same rounding/nudge approach as compute_shares)."""
    cat_sums = {cat: sum(uo_totals[u] for u in uos) for cat, uos in categories.items()}
    total = sum(cat_sums.values())
    shares = {cat: v / total for cat, v in cat_sums.items()}
    rounded = {cat: round(v, 6) for cat, v in shares.items()}
    diff = round(1.0 - sum(rounded.values()), 6)
    if diff != 0:
        largest_cat = max(rounded, key=rounded.get)
        rounded[largest_cat] = round(rounded[largest_cat] + diff, 6)
    return rounded, cat_sums, total


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    kommunal_rates, landsting_rates, sv_names = load_tax_rates()
    kommun_names = load_kommun_names()
    region_names = load_region_names()

    kommun_table, _, _ = load_jsonstat(RAW / "tab4199_kommun_2024.json")
    region_table, _, _ = load_jsonstat(RAW / "tab4242_region_2024.json")

    kommun_shares, kommun_totals = compute_shares(
        kommun_table, KOMMUN_CATEGORIES,
        exclude_codes_by_region={GOTLAND_KOMMUN_CODE: GOTLAND_EXCLUDE_CODES},
    )
    region_shares, region_totals = compute_shares(region_table, REGION_CATEGORIES)

    kommun_detail = compute_detail(
        kommun_table, KOMMUN_CATEGORY_DETAIL_CODES, KOMMUN_CODE_LABELS, kommun_totals,
        exclude_codes_by_region={GOTLAND_KOMMUN_CODE: GOTLAND_EXCLUDE_CODES},
    )
    region_detail = compute_detail(region_table, REGION_CATEGORY_DETAIL_CODES, REGION_CODE_LABELS, region_totals)

    # National (Riket) fallback shares
    riket_kommun_shares = kommun_shares.get("00")
    riket_region_shares = region_shares.get("00")
    riket_kommun_detail = kommun_detail.get("00", {})
    riket_region_detail = region_detail.get("00", {})
    assert riket_kommun_shares is not None
    assert riket_region_shares is not None

    # --- Build regions dict ---
    regions_out = {}
    for lan_code, name in sorted(region_names.items()):
        rate = landsting_rates.get(lan_code)
        assert rate is not None, f"missing landstingsskatt for lan {lan_code}"

        # find the region_table code matching this lan_code (e.g. "01L" or "0980L")
        matched_code = None
        for code in region_table.keys():
            if code == "00":
                continue
            m = re.match(r"^(\d+)L$", code)
            if m and m.group(1)[:2] == lan_code:
                matched_code = code
                break
        entry = {
            "name": name,
            "taxRate": rate,
        }
        shares = region_shares.get(matched_code) if matched_code else None
        if shares is None:
            entry["spendingShares"] = riket_region_shares
            entry["spendingDetail"] = riket_region_detail
            entry["estimated"] = True
        else:
            entry["spendingShares"] = shares
            entry["spendingDetail"] = region_detail.get(matched_code, {})
        regions_out[lan_code] = entry

    assert len(regions_out) == 21

    # --- Build kommuner dict ---
    kommuner_out = {}
    missing_spending = []
    for kod, rate in kommunal_rates.items():
        lan_code = kod[:2]
        name = kommun_names.get(kod, sv_names.get(kod, "").title())
        entry = {
            "name": name,
            "regionCode": lan_code,
            "taxRate": rate,
        }
        shares = kommun_shares.get(kod)
        if shares is None:
            entry["spendingShares"] = riket_kommun_shares
            entry["spendingDetail"] = riket_kommun_detail
            entry["estimated"] = True
            missing_spending.append(kod)
        else:
            entry["spendingShares"] = shares
            entry["spendingDetail"] = kommun_detail.get(kod, {})
        kommuner_out[kod] = entry

    assert len(kommuner_out) == 290

    # ---------------------------------------------------------------------
    # National tax parameters (2026) - see SOURCES.md for exact citations
    # ---------------------------------------------------------------------
    PBB_2026 = 59200

    national = {
        "stateTaxRate": 0.20,
        "skiktgrans": 643000,
        "prisbasbelopp": PBB_2026,
        "grundavdrag": {
            "description": (
                "Grundavdrag for personer som inte fyllt 66 ar vid inkomstarets ingang, "
                "inkomstar 2026. Piecewise-linear function of fastställd förvärvsinkomst (FFI), "
                "expressed in fractions of prisbasbeloppet (PBB). Result is capped at FFI and "
                "rounded UP to the nearest 100 kr. Source: Skatteverket SKV 433 Teknisk "
                "beskrivning 2026, avsnitt 6 'Grundavdrag'."
            ),
            "prisbasbelopp": PBB_2026,
            "roundingKr": 100,
            "roundingDirection": "up",
            "cappedAtIncome": True,
            "segments": [
                {
                    "fromFraction": 0, "toFraction": 0.99,
                    "fromKr": 0, "toKr": 58608,
                    "type": "flat", "valueFraction": 0.423, "valueKr": 25041.60,
                    "formula": "grundavdrag = 0.423 * PBB",
                },
                {
                    "fromFraction": 0.99, "toFraction": 2.72,
                    "fromKr": 58608, "toKr": 161024,
                    "type": "linear",
                    "baseFraction": 0.423, "baseKr": 25041.60,
                    "slope": 0.20,
                    "formula": "grundavdrag = 0.423*PBB + 0.20 * (FFI - 0.99*PBB)",
                },
                {
                    "fromFraction": 2.72, "toFraction": 3.11,
                    "fromKr": 161024, "toKr": 184112,
                    "type": "flat", "valueFraction": 0.77, "valueKr": 45584.00,
                    "formula": "grundavdrag = 0.77 * PBB",
                },
                {
                    "fromFraction": 3.11, "toFraction": 7.88,
                    "fromKr": 184112, "toKr": 466496,
                    "type": "linear",
                    "baseFraction": 0.77, "baseKr": 45584.00,
                    "slope": -0.10,
                    "formula": "grundavdrag = 0.77*PBB - 0.10 * (FFI - 3.11*PBB)",
                },
                {
                    "fromFraction": 7.88, "toFraction": None,
                    "fromKr": 466496, "toKr": None,
                    "type": "flat", "valueFraction": 0.293, "valueKr": 17345.60,
                    "formula": "grundavdrag = 0.293 * PBB",
                },
            ],
        },
        "jobbskatteavdrag": {
            "description": (
                "Skattereduktion for arbetsinkomst (jobbskatteavdrag), inkomstar 2026, for "
                "personer som INTE fyllt 66 ar vid inkomstarets ingang. Reduces municipal "
                "income tax ONLY (not state or regional tax). AI (arbetsinkomst) is rounded "
                "DOWN to nearest 100 kr before use; GA is the grundavdrag amount (rounded, "
                "see 'grundavdrag' object above); KI is the person's kommunal tax rate as a "
                "fraction (e.g. 0.2118), i.e. the pure kommunal rate, NOT including "
                "begravningsavgift or kyrkoavgift. Final result rounded DOWN to whole kr. "
                "Source: Skatteverket SKV 433 Teknisk beskrivning 2026, avsnitt 7.5.2 "
                "'Skattereduktion for arbetsinkomst'."
            ),
            "prisbasbelopp": PBB_2026,
            "appliesTo": "kommunalSkattOnly",
            "arbetsinkomstRounding": {"toKr": 100, "direction": "down"},
            "resultRounding": {"toKr": 1, "direction": "down"},
            "under66": {
                "segments": [
                    {
                        "fromFraction": 0, "toFraction": 0.91,
                        "fromKr": 0, "toKr": 53872,
                        "formula": "reduction = (AI - GA) * KI",
                    },
                    {
                        "fromFraction": 0.91, "toFraction": 3.24,
                        "fromKr": 53872, "toKr": 191808,
                        "baseFraction": 0.91, "baseKr": 53872, "slope": 0.3874,
                        "formula": "reduction = (0.91*PBB + 0.3874*(AI - 0.91*PBB) - GA) * KI",
                    },
                    {
                        "fromFraction": 3.24, "toFraction": 8.08,
                        "fromKr": 191808, "toKr": 478336,
                        "baseFraction": 1.813, "baseKr": 107329.60, "slope": 0.251,
                        "formula": "reduction = (1.813*PBB + 0.251*(AI - 3.24*PBB) - GA) * KI",
                    },
                    {
                        "fromFraction": 8.08, "toFraction": None,
                        "fromKr": 478336, "toKr": None,
                        "baseFraction": 3.027, "baseKr": 179198.40,
                        "formula": "reduction = (3.027*PBB - GA) * KI",
                    },
                ],
            },
            "age66Plus": {
                "note": (
                    "Simplified/different formula for persons who turned 66 before the start "
                    "of the income year; does NOT subtract grundavdrag and does NOT multiply "
                    "by KI (the kommunal tax rate) - it is a flat percentage of arbetsinkomst. "
                    "Provided for completeness; the site's default calculator targets "
                    "under-66 earners."
                ),
                "segments": [
                    {
                        "fromFraction": 0, "toFraction": 1.75,
                        "fromKr": 0, "toKr": 103600,
                        "formula": "reduction = AI * 0.22",
                    },
                    {
                        "fromFraction": 1.75, "toFraction": 5.24,
                        "fromKr": 103600, "toKr": 310208,
                        "baseFraction": 0.2635, "baseKr": 15599.20,
                        "formula": "reduction = 0.2635*PBB + 0.07*AI",
                    },
                    {
                        "fromFraction": 5.24, "toFraction": None,
                        "fromKr": 310208, "toKr": None,
                        "valueFraction": 0.6293, "valueKr": 37254.56,
                        "formula": "reduction = 0.6293*PBB (flat)",
                    },
                ],
            },
        },
    }

    categories_kommun = list(KOMMUN_CATEGORIES.keys())
    categories_region = list(REGION_CATEGORIES.keys())
    categories_state = list(STATE_CATEGORIES.keys())

    # --- Statens budget: outcome 2024 + budget 2026 by utgiftsomrade -> shares ---
    state_outcome_uo = load_state_outcome_by_uo(2024)
    state_budget_uo = load_state_budget_by_uo()
    state_outcome_shares, state_outcome_cat_sums, state_outcome_total = compute_state_shares(
        state_outcome_uo, STATE_CATEGORIES
    )
    state_budget_shares, state_budget_cat_sums, state_budget_total = compute_state_shares(
        state_budget_uo, STATE_CATEGORIES
    )
    uo_names = load_uo_names()
    state_outcome_detail = compute_state_detail(state_outcome_uo, STATE_CATEGORIES, uo_names, state_outcome_total)
    state_budget_detail = compute_state_detail(state_budget_uo, STATE_CATEGORIES, uo_names, state_budget_total)
    state_out = {
        "spendingShares": state_outcome_shares,
        "spendingSharesBudget": state_budget_shares,
        "spendingDetail": state_outcome_detail,
        "spendingDetailBudget": state_budget_detail,
    }

    output = {
        "meta": {
            "generatedAt": "2026-08-31",
            "taxRatesYear": 2026,
            "outcomeYear": 2024,
            "budgetYear": 2026,
            "budgetDataIsRealBudget": False,
            "budgetNote": (
                "No structured, cross-kommun machine-readable dataset of adopted CURRENT-YEAR "
                "budgeted spending-by-verksamhetsomrade exists at SCB or elsewhere for all ~290 "
                "kommuner and 21 regioner (budgets are published individually, mostly as PDFs, "
                "per kommun/region). The 'budget' toggle in the frontend therefore reuses these "
                "same outcomeYear (2024) actual-outcome shares as an ESTIMATE/proxy and must be "
                "labeled to the end user as such - see SOURCES.md."
            ),
            "categoriesKommun": categories_kommun,
            "categoriesRegion": categories_region,
            "categoriesState": categories_state,
            "stateBudgetDataIsRealBudget": True,
            "stateBudgetNote": (
                "Unlike the kommun/region case, the central government has exactly one budget, "
                "published in structured form. spendingSharesBudget for 'state' is real "
                "budget-year (2026) data from the government's own "
                "'Specifikation av budgetens utgifter for 2026' spreadsheet (prop. 2025/26:1), "
                "not a reused-outcome-year proxy - see SOURCES.md."
            ),
        },
        "national": national,
        "regions": regions_out,
        "kommuner": kommuner_out,
        "state": state_out,
    }

    out_path = BASE / "tax-data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, sort_keys=False)

    print(f"Wrote {out_path} ({out_path.stat().st_size:,} bytes)")
    print(f"Kommuner with estimated/fallback spending shares ({len(missing_spending)}): {missing_spending}")

    # sanity checks
    for kod, entry in kommuner_out.items():
        s = sum(entry["spendingShares"].values())
        assert abs(s - 1.0) < 1e-3, f"kommun {kod} shares sum to {s}"
        combined = entry["taxRate"] + regions_out[entry["regionCode"]]["taxRate"]
        # Gotland (0980) reports its combined kommun+region rate entirely under "kommunal"
        # (landsting portion = 0 for lan 09), so its kommunal rate alone is ~33-34%.
        if kod == "0980":
            assert 0.30 <= entry["taxRate"] <= 0.38, f"Gotland suspicious combined-as-kommunal rate {entry['taxRate']}"
        else:
            assert 0.15 <= entry["taxRate"] <= 0.25, f"kommun {kod} suspicious taxRate {entry['taxRate']}"
        assert 0.28 <= combined <= 0.38, f"kommun {kod} suspicious combined taxRate {combined}"
        # Detail sub-items for a category should sum back to that category's own
        # top-level share. Each leaf is floored at 0 individually (a negative row
        # can't be shown in the drill-down table), so wherever exactly one leaf in
        # a category went net-negative that year, dropping it makes the *visible*
        # positive leaves sum to slightly MORE than the category total, which
        # already absorbed that same negative value - the same fee-financed-activity
        # phenomenon already documented for the top-level "Övrigt" floor (see
        # compute_shares), just now also visible at the finer leaf level. Observed
        # max gap in the 2024 data is ~0.6 percentage points.
        for cat, items in entry["spendingDetail"].items():
            detail_sum = sum(i["share"] for i in items)
            assert detail_sum <= entry["spendingShares"][cat] + 1e-2, (
                f"kommun {kod} detail for {cat} ({detail_sum}) exceeds category share "
                f"({entry['spendingShares'][cat]})"
            )
    for rc, entry in regions_out.items():
        s = sum(entry["spendingShares"].values())
        assert abs(s - 1.0) < 1e-3, f"region {rc} shares sum to {s}"
        if rc == "09":
            # Gotland: the region/landsting rate is folded into the kommun's rate (see above).
            assert entry["taxRate"] == 0.0, f"Gotland region rate expected 0, got {entry['taxRate']}"
        else:
            assert 0.09 <= entry["taxRate"] <= 0.135, f"region {rc} suspicious taxRate {entry['taxRate']}"
        for cat, items in entry["spendingDetail"].items():
            detail_sum = sum(i["share"] for i in items)
            assert detail_sum <= entry["spendingShares"][cat] + 1e-2, (
                f"region {rc} detail for {cat} ({detail_sum}) exceeds category share "
                f"({entry['spendingShares'][cat]})"
            )

    s_out = sum(state_out["spendingShares"].values())
    s_bud = sum(state_out["spendingSharesBudget"].values())
    assert abs(s_out - 1.0) < 1e-6, f"state outcome shares sum to {s_out}"
    assert abs(s_bud - 1.0) < 1e-6, f"state budget shares sum to {s_bud}"
    for cat in categories_state:
        d_out = sum(i["share"] for i in state_outcome_detail[cat])
        d_bud = sum(i["share"] for i in state_budget_detail[cat])
        # 1e-5 tolerance: compute_state_shares nudges the largest category so its
        # rounded shares sum to exactly 1.0, which can leave a ~1e-6 gap against
        # the independently-rounded per-UO detail shares - benign rounding noise.
        assert abs(d_out - state_outcome_shares[cat]) < 1e-5, f"state outcome detail mismatch for {cat}"
        assert abs(d_bud - state_budget_shares[cat]) < 1e-5, f"state budget detail mismatch for {cat}"
    print(f"State outcome total 2024: {state_outcome_total/1000:.1f} mdkr; "
          f"budget total 2026: {state_budget_total/1000:.1f} mdkr")
    for cat in categories_state:
        print(f"  {cat:45s} outcome {state_outcome_shares[cat]*100:5.2f}%   "
              f"budget {state_budget_shares[cat]*100:5.2f}%")

    print("All sanity checks passed.")


if __name__ == "__main__":
    main()
